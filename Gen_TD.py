import os
import sys
import logging
import settings as s

# Logging settings
logging.basicConfig(filename='log.log', level=logging.DEBUG,
                    format='%(asctime)s:%(levelname)s:%(message)s')
logging.info('START')

"""Try to import openpyxl, otherwise prompt user"""
try:
    import openpyxl as xl
except ModuleNotFoundError:
    print('Error! openpyxl module not found, please install it')
    logging.error('openpyxl module not found, please install it')
    sys.exit()

# Version 0.92


# Open excel file
try:
    wb = xl.load_workbook(s.EXCEL_FILE, data_only=True)
except FileNotFoundError as e:
    print(e)
    logging.error(e)
    print('Error! Excel file not found, program will exit')
    logging.error('Excel file not found, program will exit')
    sys.exit()


# Funtions
def td_single(config_file, ref_txt):
    """Read a text file and copy the data inside notifiers to memory"""
    with open(config_file, 'r') as config:
        exists_in_config = False
        section_found = False
        inst_data = ''
        begin = '[gen.' + ref_txt + '_begin]'
        end = '[gen.' + ref_txt + '_end]'

        for line in config:
            if end in str(line):
                section_found = False
            if section_found:
                inst_data += line
            if begin in str(line):
                exists_in_config = True
                section_found = True
    if not exists_in_config:
        print(ref_txt, 'not found in config file!')
        logging.warning(ref_txt + ' not found in config file!')

    return inst_data


def td_multiple(config_file, ref_txt, excelsheet):
    """Get text lines from config file and replace by data in excel, then append the new lines to memory"""

    # Check if sheet exists
    sheets = wb.sheetnames
    for sheet in sheets:
        if sheet == excelsheet:
            sheet_exists = True
            break
        else:
            sheet_exists = False
    if sheet_exists:
        active_sheet = wb[excelsheet]  # Open sheet
    else:
        print('Error!', excelsheet, "Sheet doesn't exist!", "program will exit")
        logging.error(excelsheet + " sheet doesn't exist!" + " program will exit")
        sys.exit()

    if sheet_exists:
        with open(config_file, 'r') as config:
            exists_in_config = False
            section_found = False
            inst_data = ''
            begin = '[gen.' + ref_txt + '_begin]'
            end = '[gen.' + ref_txt + '_end]'
            zero_index = 0

            for i in range(s.ROW, active_sheet.max_row + 1):
                cell_id = active_sheet.cell(row=i, column=s.COL_ID)
                cell_config = active_sheet.cell(row=i, column=s.COL_CONFIG)
                cell_comment = active_sheet.cell(row=i, column=s.COL_COMMENT)
                zero_index += 1

                if cell_id.value is None:
                    break

                config.seek(0, 0)  # Seek to beginning of file
                for index, line in enumerate(config, start=1):
                    if end in str(line):
                        section_found = False
                    if section_found:
                        line = line.replace(s.ID_REPLACE, cell_id.value)

                        if cell_config.value is not None:
                            line = line.replace(s.CONFIG_REPLACE, cell_config.value)
                        line = line.replace(s.INDEX_REPLACE, str(zero_index))

                        # check if comment exists, if insert empty string
                        if cell_comment.value is None:
                            line = line.replace(s.COMMENT_REPLACE, '')
                        else:
                            line = line.replace(s.COMMENT_REPLACE, cell_comment.value)

                        inst_data += line  # Create instance data
                    if begin in str(line):
                        exists_in_config = True
                        section_found = True
            if not exists_in_config:
                print(ref_txt, 'in config file not found!')
                logging.warning(ref_txt + ' in config file not found!')
            else:
                print('Generated from row:', s.ROW, 'to', i - 1, 'of', ref_txt, 'in', active_sheet)
                logging.info(
                    'Generated from row: ' + str(s.ROW) + ' to ' + str(i - 1) + ' of ' + ref_txt + 'in' + str(active_sheet))

        return inst_data


def td_multiple_config(sub_dir, ref_txt, excelsheet):
    """Same as td_multiple, but config stored in different files"""
    # Check if sheet exists
    sheets = wb.sheetnames
    for sheet in sheets:
        if sheet == excelsheet:
            sheet_exists = True
            break
        else:
            sheet_exists = False
    if sheet_exists:
        active_sheet = wb[excelsheet]  # Open sheet
    else:
        print('Error!', excelsheet, "Sheet doesn't exist!", "program will exit")
        logging.error(excelsheet + " sheet doesn't exist!" + " program will exit")
        sys.exit()

    if sheet_exists:
        # Setup variables
        exists_in_config = False
        section_found = False
        inst_data = ''
        begin = '[gen.' + ref_txt + '_begin]'
        end = '[gen.' + ref_txt + '_end]'
        zero_index = 0

        # loop through excel rows, get value at corresponding cell
        for i in range(s.ROW, active_sheet.max_row + 1):
            cell_id = active_sheet.cell(row=i, column=s.COL_ID)
            cell_config = active_sheet.cell(row=i, column=s.COL_CONFIG)
            cell_comment = active_sheet.cell(row=i, column=s.COL_COMMENT)
            zero_index += 1

            if cell_id.value is None:
                break

            # combine file path and open corresponding file
            filename = cell_config.value + '.txt'
            file_and_path = os.path.join(sub_dir, filename)
            with open(file_and_path, 'r') as config:
                for line in config:
                    if end in str(line):
                        section_found = False
                    if section_found:
                        line = line.replace(s.ID_REPLACE, cell_id.value)

                        if cell_config.value is not None:
                            line = line.replace(s.CONFIG_REPLACE, cell_config.value)
                        line = line.replace(s.INDEX_REPLACE, str(zero_index))

                        # check if comment exists, if insert empty string
                        if cell_comment.value is None:
                            line = line.replace(s.COMMENT_REPLACE, '')
                        else:
                            line = line.replace(s.COMMENT_REPLACE, cell_comment.value)

                        inst_data += line  # Create instance data
                    if begin in str(line):
                        exists_in_config = True
                        section_found = True
        if not exists_in_config:
            print(ref_txt, 'in config file not found!')
            logging.warning(ref_txt + ' in config file not found!')
        else:
            print('Generated from row:', s.ROW, 'to', i - 1, 'of', ref_txt, 'in', active_sheet)
            logging.info(
                'Generated from row: ' + str(s.ROW) + ' to ' + str(i - 1) + ' of ' + ref_txt + 'in' + str(active_sheet))

        return inst_data


# DI function
def td_gen_di():
    """Create and concetenate all text lines to different files"""
    # setup variables
    config_file = os.path.join(s.CONFIG_PATH, 'Config_DI.txt')
    sheet = 'DI'

    # Check what output path to use, if 'None' create in current directory, otherwise as specified
    if s.OUTPUT_PATH is None:
        file_path = 'Generated DI'
    else:
        file_path = os.path.join(s.OUTPUT_PATH, 'Generated DI')
    # Create sub-directory if it doesn't exist
    if not os.path.exists(file_path):
        os.makedirs(file_path)

    # PLC function, concatenate data
    header_data = td_single(config_file, 'header')
    var_data = td_multiple(config_file, 'var', sheet)
    func_header_data = td_single(config_file, 'funcHeader')
    codebody_data = td_multiple(config_file, 'codebody', sheet)
    footer_data = td_single(config_file, 'footer')

    # Create file and put it inside path created above
    filename = 'PLC_' + sheet + '.awl'
    file_and_path = os.path.join(file_path, filename)
    with open(file_and_path, 'w') as functionFile:
        data = header_data
        data += var_data
        data += func_header_data
        data += codebody_data
        data += footer_data
        functionFile.write(data)
        print(filename, 'created')
        logging.info(filename + ' created')

    # PLC Datablock, if all elements exists concatenate data and create file
    db_header_data = td_single(config_file, 'db_header')
    db_var_data = td_multiple(config_file, 'db_var', sheet)
    db_footer_data = td_single(config_file, 'db_footer')
    if db_header_data != '' and db_var_data != '' and db_footer_data != '':
        filename = 'PLC_' + sheet + '_DB.db'
        file_and_path = os.path.join(file_path, filename)
        with open(file_and_path, 'w') as dbFile:
            data = db_header_data
            data += db_var_data
            data += db_footer_data
            dbFile.write(data)
            print(filename, 'created')
            logging.info(filename + ' created')

    # PLC symbol table
    symbol_data = td_multiple(config_file, 'symbol', sheet)
    if symbol_data != '':
        filename = 'PLC_' + sheet + '_Symbol.sdf'
        file_and_path = os.path.join(file_path, filename)
        with open(file_and_path, 'w') as symbolFile:
            symbolFile.write(symbol_data)
            print(filename, 'created')
            logging.info(filename + ' created')

    # Intouch
    it_data = td_multiple(config_file, 'Intouch', sheet)
    if it_data != '':
        filename = 'IT_' + sheet + '.csv'
        file_and_path = os.path.join(file_path, filename)
        with open(file_and_path, 'w') as itFile:
            itFile.write(it_data)
            print(filename, 'created')
            logging.info(filename + ' created')
    print('Generated files put in...', file_path)
    logging.info('Generated DI files put in ' + file_path)


# DO function
def td_gen_do():
    """Create and concetenate all text lines to different files"""
    # setup variables
    config_file = os.path.join(s.CONFIG_PATH, 'Config_DO.txt')
    sheet = 'DO'

    # Check what output path to use, if 'None' create in current directory, otherwise as specified
    if s.OUTPUT_PATH is None:
        file_path = 'Generated DO'
    else:
        file_path = os.path.join(s.OUTPUT_PATH, 'Generated DO')
    # Create sub-directory if it doesn't exist
    if not os.path.exists(file_path):
        os.makedirs(file_path)

    # PLC function, concatenate data
    header_data = td_single(config_file, 'header')
    var_data = td_multiple(config_file, 'var', sheet)
    func_header_data = td_single(config_file, 'funcHeader')
    codebody_data = td_multiple(config_file, 'codebody', sheet)
    footer_data = td_single(config_file, 'footer')

    # Create file and put it inside path created above
    filename = 'PLC_' + sheet + '.awl'
    file_and_path = os.path.join(file_path, filename)
    with open(file_and_path, 'w') as functionFile:
        data = header_data
        data += var_data
        data += func_header_data
        data += codebody_data
        data += footer_data
        functionFile.write(data)
        print(filename, 'created')
        logging.info(filename + ' created')

    # PLC Datablock, if all elements exists concatenate data and create file
    db_header_data = td_single(config_file, 'db_header')
    db_var_data = td_multiple(config_file, 'db_var', sheet)
    db_footer_data = td_single(config_file, 'db_footer')
    if db_header_data != '' and db_var_data != '' and db_footer_data != '':
        filename = 'PLC_' + sheet + '_DB.db'
        file_and_path = os.path.join(file_path, filename)
        with open(file_and_path, 'w') as dbFile:
            data = db_header_data
            data += db_var_data
            data += db_footer_data
            dbFile.write(data)
            print(filename, 'created')
            logging.info(filename + ' created')

    # PLC symbol table
    symbol_data = td_multiple(config_file, 'symbol', sheet)
    if symbol_data != '':
        filename = 'PLC_' + sheet + '_Symbol.sdf'
        file_and_path = os.path.join(file_path, filename)
        with open(file_and_path, 'w') as symbolFile:
            symbolFile.write(symbol_data)
            print(filename, 'created')
            logging.info(filename + ' created')

    # Intouch
    it_data = td_multiple(config_file, 'Intouch', sheet)
    if it_data != '':
        filename = 'IT_' + sheet + '.csv'
        file_and_path = os.path.join(file_path, filename)
        with open(file_and_path, 'w') as itFile:
            itFile.write(it_data)
            print(filename, 'created')
            logging.info(filename + ' created')
    print('Generated files put in...', file_path)
    logging.info('Generated DO files put in ' + file_path)


# Valve function
def td_gen_valve():
    """Create and concetenate all text lines to different files"""
    # TODO rest of the code
    # setup variables
    config_file = os.path.join(s.CONFIG_PATH_VALVE, 'Config_valve.txt')
    sheet = 'Valves'

    # Check what output path to use, if 'None' create in current directory, otherwise as specified
    if s.OUTPUT_PATH is None:
        file_path = 'Generated Valve'
    else:
        file_path = os.path.join(s.OUTPUT_PATH, 'Generated Valve')
    # Create sub-directory if it doesn't exist
    if not os.path.exists(file_path):
        os.makedirs(file_path)

    # PLC function, concatenate data
    codebody_data = td_multiple_config(s.CONFIG_PATH_VALVE, 'codebody', sheet)

    # Create file and put it inside path created above
    filename = 'PLC_' + sheet + '.awl'
    file_and_path = os.path.join(file_path, filename)
    with open(file_and_path, 'w') as functionFile:
        data = codebody_data

        functionFile.write(data)
        print(filename, 'created')
        logging.info(filename + ' created')

    print('Generated files put in...', file_path)
    logging.info('Generated DO files put in ' + file_path)


"""Call functions, don't execute if disabled in settings.py"""
# DI
if s.DI_DISABLE:
    print('DI not generated, disabled in settings file')
    logging.warning('DI not generated, disabled in settings file')
else:
    td_gen_di()

# DO
if s.DO_DISABLE:
    print('DO not generated, disabled in settings file')
    logging.warning('DO not generated, Disabled in settings file')
else:
    td_gen_do()

# Valve
if s.VALVE_DISABLE:
    print('Valve not generated, disabled in settings file')
    logging.warning('Valve not generated, Disabled in settings file')
else:
    td_gen_valve()

logging.info('STOP')
