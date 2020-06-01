import sys
import logging
import os
import os.path
from os import path
import tkinter as tk
from tkinter import filedialog
import pickle
import settings as s

"""Try to import openpyxl, otherwise prompt user"""
try:
    import openpyxl as xl
except ModuleNotFoundError:
    print('Error! openpyxl module not found, please install it')
    logging.error('openpyxl module not found, please install it')
    sys.exit()

"""Version variable"""
version = ('Version', str(0.97))
print(version)

"""Load user data if it exists, otherwise create dictionary"""
EXCEL_PATH_START_VALUE = 'No excel specified'
OUTPUT_PATH_START_VALUE = 'No path specified'
if path.exists('user_data.pickle'):
    with open('user_data.pickle', 'rb') as f:
        user_data = pickle.load(f)
else:
    # if file doesn't exist initialize data to start-values
    user_data = {
        'excel_path': EXCEL_PATH_START_VALUE,
        'output_path': OUTPUT_PATH_START_VALUE,
    }


class TdUI(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack()

        # Constants
        self.height = 200
        self.width = 500
        self.frameColor = "#2b2b2b"
        self.buttonWidth = 0.2
        self.buttonHeight = 0.12
        self.button_bg = "#2b2b2b"
        self.button_fg = "#FFFFFF"  # text color
        self.fontSize = 10
        self.buttonYSpacing = 0.20
        self.checkbuttonYSpacing = 0.1

        # Call app
        self.create_window()
        self.create_window_contents()
        self.create_dropdown()
        self.check_path_validity()  # "Run Script" button changes state from this function

    def create_window(self):
        """Create window"""
        # Title and program-icon
        self.master.title('MC TD Gen')
        # self.master.iconbitmap('C:\') TODO Program Icon

        self.canvas = tk.Canvas(self.master, height=self.height, width=self.width)
        self.canvas.pack()
        self.frame = tk.Frame(self.master, bg=self.frameColor)
        self.frame.place(relwidth=1, relheight=1)

    def create_dropdown(self):
        """Create drop-down menu"""
        self.menu = tk.Menu(self.master)
        self.master.config(menu=self.menu)

        # file submenu
        self.subMenu = tk.Menu(self.menu, tearoff=0)
        self.menu.add_cascade(label="File", menu=self.subMenu)
        self.subMenu.add_command(label="Exit", command=self.master.quit)

        # view submenu
        self.viewSubMenu = tk.Menu(self.menu, tearoff=0)
        self.menu.add_cascade(label="View", menu=self.viewSubMenu)
        self.viewSubMenu.add_command(label="Log file", command=self.open_logfile)
        self.viewSubMenu.add_command(label="Settings file", command=self.open_settings)
        self.viewSubMenu.add_command(label="Config files", command=self.open_config_path)

        # about submenu
        self.aboutSubMenu = tk.Menu(self.menu, tearoff=0)
        self.menu.add_cascade(label="About", menu=self.aboutSubMenu)
        self.aboutSubMenu.add_command(label="Version", command=self.create_about_window)

    def create_window_contents(self):
        """Create window contents"""
        # Excel button
        self.excelButton = tk.Button(self.master, text="Select Excel...", bg=self.button_bg, fg=self.button_fg,
                                     command=self.browse_excel)
        self.excelButton.place(relx=0.03, rely=0.1, relheight=self.buttonHeight, relwidth=self.buttonWidth)
        # Excel path label
        self.excelLabel = tk.Label(self.frame, bg=self.button_bg, fg=self.button_fg, text=(user_data['excel_path']))
        self.excelLabel.place(relx=0.25, rely=0.1, relheight=self.buttonHeight)

        # Output path button
        self.outpathButton = tk.Button(self.master, text="Output path...", bg=self.button_bg, fg=self.button_fg,
                                       command=self.output_path)
        self.outpathButton.place(relx=0.03, rely=0.1 + self.buttonYSpacing, relheight=self.buttonHeight,
                                 relwidth=self.buttonWidth)
        # Output path label
        self.outpathLabel = tk.Label(self.frame, bg=self.button_bg, fg=self.button_fg, text=(user_data['output_path']))
        self.outpathLabel.place(relx=0.25, rely=0.1 + self.buttonYSpacing, relheight=self.buttonHeight)

        # Run script
        self.run_self = tk.Button(self.master, text="Run script", bg=self.button_bg, fg=self.button_fg,
                                  command=self.run_self, state=tk.DISABLED)
        self.run_self.place(relx=0.03, rely=0.75, relheight=self.buttonHeight, relwidth=self.buttonWidth)

    def browse_excel(self):
        excelPath = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
        # Write to user_data dictionary, to save it for later.
        user_data['excel_path'] = excelPath
        # Update label
        self.excelLabel.config(text=excelPath)

        # Check if all path are valid
        self.check_path_validity()

    def output_path(self):
        output_path = filedialog.askdirectory()
        # Write to user_data dictionary, to save it for later.
        user_data['output_path'] = output_path
        # Update label
        self.outpathLabel.config(text=output_path)

    def open_logfile(self):
        os.system('log.log')

    def open_settings(self):
        os.system('settings.py')

    def run_python_windows_command(self):
        os.system('run.bat')

    def check_path_validity(self):
        if os.path.isfile(user_data['excel_path']):
            self.run_self.configure(state=tk.NORMAL)
        else:
            self.run_self.configure(state=tk.DISABLED)

    def run_self(self):
        GenTD(user_data['excel_path'], user_data['output_path'])

    def open_config_path(self):
        path = s.CONFIG_PATH
        path = os.path.realpath(path)
        os.startfile(path)

    def create_about_window(self):
        self.about = tk.Toplevel()
        self.about.title('About')
        # self.about.iconbitmap('C:\') TODO Program Icon
        self.label = tk.Label(self.about, text=version).pack()


class GenTD:
    def __init__(self, excel_path, output_path):
        self.excel_path = excel_path
        self.output_path = output_path

        self.generate()

    def generate(self):
        """Logging settings"""
        logging.basicConfig(filename='log.log', level=logging.DEBUG,
                            format='%(asctime)s:%(levelname)s:%(message)s')
        logging.info('START')

        self.open_td_excel()

        # DI
        if s.DI_DISABLE:
            print('DI not generated, disabled in settings file')
            logging.warning('DI not generated, disabled in settings file')
        else:
            self.td_gen_di()

        # DO
        if s.DO_DISABLE:
            print('DO not generated, disabled in settings file')
            logging.warning('DO not generated, Disabled in settings file')
        else:
            self.td_gen_do()

        # Valve
        if s.VALVE_DISABLE:
            print('Valve not generated, disabled in settings file')
            logging.warning('Valve not generated, Disabled in settings file')
        else:
            self.td_gen_valve()

        # Motor
        if s.MOTOR_DISABLE:
            print('Motor not generated, disabled in settings file')
            logging.warning('Motor not generated, Disabled in settings file')
        else:
            self.td_gen_motor()

        # AI
        if s.AI_DISABLE:
            print('AI not generated, disabled in settings file')
            logging.warning('AI not generated, Disabled in settings file')
        else:
            self.td_gen_AI()

        logging.info('STOP')

    def open_td_excel(self):
        try:
            wb = xl.load_workbook(self.excel_path, data_only=True)
        except FileNotFoundError as e:
            print(e)
            logging.error(e)
            print('Error! Excel file not found, program will exit')
            logging.error('Excel file not found, program will exit')
            sys.exit()
        else:
            self.wb = wb

    def td_single(self, config_file, ref_txt):
        """Read a text file and copy the data inside notifiers to memory"""
        with open(config_file, 'r') as config:
            exists_in_config = False
            section_found = False
            inst_data = ''
            begin = '[gen.' + ref_txt + '_begin]'
            end = '[gen.' + ref_txt + '_end]'

            for line in config:
                if end in str(line):
                    section_found = False
                if section_found:
                    inst_data += line
                if begin in str(line):
                    exists_in_config = True
                    section_found = True
        if not exists_in_config:
            print(ref_txt, 'not found in config file!')
            logging.warning(ref_txt + ' not found in config file!')

        return inst_data

    def td_multiple(self, config_file, ref_txt, excelsheet, udt_size=30, udt_offset=14, start_index=0):
        """Get text lines from config file and replace by data in excel, then append the new lines to memory"""
        # Try to open excelsheet, otherwise prompt user
        try:
            active_sheet = self.wb[excelsheet]  # open sheet
        except KeyError:
            msg = f'Error! {excelsheet} sheet does not exist, program will exit'
            print(msg)
            logging.error(msg)
            sys.exit()

        with open(config_file, 'r') as config:
            exists_in_config = False
            section_found = False
            inst_data = ''
            begin = '[gen.' + ref_txt + '_begin]'
            end = '[gen.' + ref_txt + '_end]'
            addressIndex = start_index - 1  # subtract 1 as the index is increased directly in loop below

            for i in range(s.ROW, active_sheet.max_row + 1):
                cell_id = active_sheet.cell(row=i, column=s.COL_ID)
                cell_config = active_sheet.cell(row=i, column=s.COL_CONFIG)
                cell_comment = active_sheet.cell(row=i, column=s.COL_COMMENT)
                cell_engunit = active_sheet.cell(row=i, column=s.COL_ENG_UNIT)
                cell_engmin = active_sheet.cell(row=i, column=s.COL_ENG_MIN)
                cell_engmax = active_sheet.cell(row=i, column=s.COL_ENG_MAX)

                addressIndex += 1

                if cell_id.value is None:
                    break

                config.seek(0, 0)  # Seek to beginning of file
                for lineIndex, line in enumerate(config, start=1):
                    if end in str(line):
                        section_found = False
                    if section_found:
                        line = line.replace(s.ID_REPLACE, cell_id.value)

                        if cell_config.value is not None:
                            line = line.replace(s.CONFIG_REPLACE, cell_config.value)

                        # Replace index
                        line = line.replace(s.INDEX_REPLACE, str(addressIndex))

                        # calculate address by offset & datatype udt_size
                        adress = (addressIndex * udt_size) + udt_offset
                        # Replace '@ADR'
                        line = line.replace(s.ADR_REPLACE, str(adress))

                        # Replace PLC
                        line = line.replace(s.PLC_REPLACE, s.PLC_NAME)

                        # check if eng unit exists, if not insert empty string
                        if cell_engunit.value is None:
                            line = line.replace(s.ENG_UNIT_REPLACE, '')
                        else:
                            line = line.replace(s.ENG_UNIT_REPLACE, cell_engunit.value)

                        # check if eng min exists, if not insert 0
                        if cell_engmin.value is None:
                            line = line.replace(s.ENG_MIN_REPLACE, '0')
                        else:
                            line = line.replace(s.ENG_MIN_REPLACE, str(cell_engmin.value))

                        # check if eng max exists, if not insert 100
                        if cell_engmax.value is None:
                            line = line.replace(s.ENG_MAX_REPLACE, '100')
                        else:
                            line = line.replace(s.ENG_MAX_REPLACE, str(cell_engmax.value))

                        # check if comment exists, if not insert empty string
                        if cell_comment.value is None:
                            line = line.replace(s.COMMENT_REPLACE, '')
                        else:
                            line = line.replace(s.COMMENT_REPLACE, cell_comment.value)

                        inst_data += line  # Create instance data
                    if begin in str(line):
                        exists_in_config = True
                        section_found = True
            if not exists_in_config:
                print(ref_txt, 'in config file not found!')
                logging.warning(ref_txt + ' in config file not found!')
            else:
                print('Generated from row:', s.ROW, 'to', i - 1, 'of', ref_txt, 'in', active_sheet)
                logging.info(
                    'Generated from row: ' + str(s.ROW) + ' to ' + str(i - 1) + ' of ' + ref_txt + 'in' + str(
                        active_sheet))

        return inst_data

    def td_multiple_config(self, sub_dir, ref_txt, excelsheet):
        """Same as td_multiple, but config stored in different files"""
        # Try to open excelsheet, otherwise prompt user
        try:
            active_sheet = self.wb[excelsheet]  # open sheet
        except KeyError:
            msg = f'Error! {excelsheet} sheet does not exist, program will exit'
            print(msg)
            logging.error(msg)
            sys.exit()

        # Setup variables
        exists_in_config = False
        section_found = False
        inst_data = ''
        begin = '[gen.' + ref_txt + '_begin]'
        end = '[gen.' + ref_txt + '_end]'
        zero_index = 0

        # loop through excel rows, get value at corresponding cell
        for i in range(s.ROW, active_sheet.max_row + 1):
            cell_id = active_sheet.cell(row=i, column=s.COL_ID)
            cell_config = active_sheet.cell(row=i, column=s.COL_CONFIG)
            cell_comment = active_sheet.cell(row=i, column=s.COL_COMMENT)
            zero_index += 1

            if cell_id.value is None:
                break

            # combine file path and open corresponding file
            filename = cell_config.value + '.txt'
            file_and_path = os.path.join(sub_dir, filename)
            with open(file_and_path, 'r') as config:
                for line in config:
                    if end in str(line):
                        section_found = False
                    if section_found:
                        line = line.replace(s.ID_REPLACE, cell_id.value)

                        if cell_config.value is not None:
                            line = line.replace(s.CONFIG_REPLACE, cell_config.value)
                        line = line.replace(s.INDEX_REPLACE, str(zero_index))

                        # check if comment exists, if insert empty string
                        if cell_comment.value is None:
                            line = line.replace(s.COMMENT_REPLACE, '')
                        else:
                            line = line.replace(s.COMMENT_REPLACE, cell_comment.value)

                        inst_data += line  # Create instance data
                    if begin in str(line):
                        exists_in_config = True
                        section_found = True
        if not exists_in_config:
            print(ref_txt, 'in config file not found!')
            logging.warning(ref_txt + ' in config file not found!')
        else:
            print('Generated from row:', s.ROW, 'to', i - 1, 'of', ref_txt, 'in', active_sheet)
            logging.info(
                'Generated from row: ' + str(s.ROW) + ' to ' + str(i - 1) + ' of ' + ref_txt + 'in' + str(active_sheet))
        return inst_data

    def td_gen_di(self):
        """Create and concetenate all text lines to different files"""
        # setup variables
        config_file = os.path.join(s.CONFIG_PATH, 'Config_DI.txt')
        sheet = 'DI'

        # Check what output path to use, if 'None' create in current directory, otherwise as specified
        if self.output_path is None:
            file_path = 'Generated DI'
        elif self.output_path == OUTPUT_PATH_START_VALUE:
            file_path = 'Generated DI'

        else:
            file_path = os.path.join(self.output_path, 'Generated DI')
        # Create sub-directory if it doesn't exist
        if not os.path.exists(file_path):
            os.makedirs(file_path)

        # PLC function, concatenate data
        header_data = self.td_single(config_file, 'header')
        var_data = self.td_multiple(config_file, 'var', sheet)
        func_header_data = self.td_single(config_file, 'funcHeader')
        codebody_data = self.td_multiple(config_file, 'codebody', sheet)
        footer_data = self.td_single(config_file, 'footer')

        # Create file and put it inside path created above
        filename = 'PLC_' + sheet + '.awl'
        file_and_path = os.path.join(file_path, filename)
        with open(file_and_path, 'w') as functionFile:
            data = header_data
            data += var_data
            data += func_header_data
            data += codebody_data
            data += footer_data
            functionFile.write(data)
            print(filename, 'created')
            logging.info(filename + ' created')

        # PLC Datablock, if all elements exists concatenate data and create file
        db_header_data = self.td_single(config_file, 'db_header')
        db_var_data = self.td_multiple(config_file, 'db_var', sheet)
        db_footer_data = self.td_single(config_file, 'db_footer')
        if db_header_data != '' and db_var_data != '' and db_footer_data != '':
            filename = 'PLC_' + sheet + '_DB.db'
            file_and_path = os.path.join(file_path, filename)
            with open(file_and_path, 'w') as dbFile:
                data = db_header_data
                data += db_var_data
                data += db_footer_data
                dbFile.write(data)
                print(filename, 'created')
                logging.info(filename + ' created')

        # PLC symbol table
        symbol_data = self.td_multiple(config_file, 'symbol', sheet)
        if symbol_data != '':
            filename = 'PLC_' + sheet + '_Symbol.sdf'
            file_and_path = os.path.join(file_path, filename)
            with open(file_and_path, 'w') as symbolFile:
                symbolFile.write(symbol_data)
                print(filename, 'created')
                logging.info(filename + ' created')

        # Intouch
        it_data = self.td_multiple(config_file, 'Intouch', sheet, start_index=s.DI_START_INDEX)
        if it_data != '':
            filename = 'IT_' + sheet + '.csv'
            file_and_path = os.path.join(file_path, filename)
            with open(file_and_path, 'w') as itFile:
                itFile.write(it_data)
                print(filename, 'created')
                logging.info(filename + ' created')
        print('Generated files put in...', file_path)
        logging.info('Generated DI files put in ' + file_path)

    def td_gen_do(self):
        """Create and concetenate all text lines to different files"""
        # setup variables
        config_file = os.path.join(s.CONFIG_PATH, 'Config_DO.txt')
        sheet = 'DO'

        # Check what output path to use, if 'None' create in current directory, otherwise as specified
        if self.output_path is None:
            file_path = 'Generated DO'
        elif self.output_path == OUTPUT_PATH_START_VALUE:
            file_path = 'Generated DO'
        else:
            file_path = os.path.join(self.output_path, 'Generated DO')
        # Create sub-directory if it doesn't exist
        if not os.path.exists(file_path):
            os.makedirs(file_path)

        # PLC function, concatenate data
        header_data = self.td_single(config_file, 'header')
        var_data = self.td_multiple(config_file, 'var', sheet)
        func_header_data = self.td_single(config_file, 'funcHeader')
        codebody_data = self.td_multiple(config_file, 'codebody', sheet)
        footer_data = self.td_single(config_file, 'footer')

        # Create file and put it inside path created above
        filename = 'PLC_' + sheet + '.awl'
        file_and_path = os.path.join(file_path, filename)
        with open(file_and_path, 'w') as functionFile:
            data = header_data
            data += var_data
            data += func_header_data
            data += codebody_data
            data += footer_data
            functionFile.write(data)
            print(filename, 'created')
            logging.info(filename + ' created')

        # PLC Datablock, if all elements exists concatenate data and create file
        db_header_data = self.td_single(config_file, 'db_header')
        db_var_data = self.td_multiple(config_file, 'db_var', sheet)
        db_footer_data = self.td_single(config_file, 'db_footer')
        if db_header_data != '' and db_var_data != '' and db_footer_data != '':
            filename = 'PLC_' + sheet + '_DB.db'
            file_and_path = os.path.join(file_path, filename)
            with open(file_and_path, 'w') as dbFile:
                data = db_header_data
                data += db_var_data
                data += db_footer_data
                dbFile.write(data)
                print(filename, 'created')
                logging.info(filename + ' created')

        # PLC symbol table
        symbol_data = self.td_multiple(config_file, 'symbol', sheet)
        if symbol_data != '':
            filename = 'PLC_' + sheet + '_Symbol.sdf'
            file_and_path = os.path.join(file_path, filename)
            with open(file_and_path, 'w') as symbolFile:
                symbolFile.write(symbol_data)
                print(filename, 'created')
                logging.info(filename + ' created')

        # Intouch
        it_data = self.td_multiple(config_file, 'Intouch', sheet, start_index=s.DO_START_INDEX)
        if it_data != '':
            filename = 'IT_' + sheet + '.csv'
            file_and_path = os.path.join(file_path, filename)
            with open(file_and_path, 'w') as itFile:
                itFile.write(it_data)
                print(filename, 'created')
                logging.info(filename + ' created')
        print('Generated files put in...', file_path)
        logging.info('Generated DO files put in ' + file_path)

    def td_gen_valve(self):
        """Create and concetenate all text lines to different files"""
        # setup variables
        config_file = os.path.join(s.CONFIG_PATH_VALVE, 'Config_valve.txt')
        sheet = 'Valves'

        # Check what output path to use, if 'None' create in current directory, otherwise as specified
        if self.output_path is None:
            file_path = 'Generated Valve'
        elif self.output_path == OUTPUT_PATH_START_VALUE:
            file_path = 'Generated Valve'
        else:
            file_path = os.path.join(self.output_path, 'Generated Valve')
        # Create sub-directory if it doesn't exist
        if not os.path.exists(file_path):
            os.makedirs(file_path)

        """Intouch IO:Int"""
        IT_IOInt_header = self.td_single(config_file, 'IT_IOInt_Header')
        IT_IOInt_data = self.td_multiple(config_file, 'IT_IOInt_Tag', sheet, udt_size=30, udt_offset=14,
                                         start_index=s.VALVE_START_INDEX)

        """Intouch Memory:Int"""
        IT_MemInt_header = self.td_single(config_file, 'IT_MemInt_Header')
        IT_MemInt_data = self.td_multiple(config_file, 'IT_MemInt_Tag', sheet, start_index=s.VALVE_START_INDEX)

        if IT_IOInt_data != '' and IT_IOInt_header != '' and IT_MemInt_header != '' and IT_MemInt_data != '':
            filename = 'IT_' + sheet + '.csv'
            file_and_path = os.path.join(file_path, filename)
            with open(file_and_path, 'w') as itFile:
                data = IT_IOInt_header
                data += IT_IOInt_data
                data += IT_MemInt_header
                data += IT_MemInt_data

                itFile.write(data)
                print(filename, 'created')
                logging.info(filename + ' created')
        print('Generated files put in...', file_path)
        logging.info('Generated Valve files put in ' + file_path)

    def td_gen_motor(self):
        """Create and concetenate all text lines to different files"""
        # setup variables
        config_file = os.path.join(s.CONFIG_PATH, 'Config_motor.txt')
        sheet = 'Motors'

        # Check what output path to use, if 'None' create in current directory, otherwise as specified
        if self.output_path is None:
            file_path = 'Generated Motor'
        elif self.output_path == OUTPUT_PATH_START_VALUE:
            file_path = 'Generated Motor'
        else:
            file_path = os.path.join(self.output_path, 'Generated Motor')
        # Create sub-directory if it doesn't exist
        if not os.path.exists(file_path):
            os.makedirs(file_path)

        """Intouch IO:Int"""
        IT_IOInt_header = self.td_single(config_file, 'IT_IOInt_Header')
        IT_IOInt_data = self.td_multiple(config_file, 'IT_IOInt_Tag', sheet, udt_size=30, udt_offset=14,
                                         start_index=s.MOTOR_START_INDEX)
        """Intouch Memory:Int"""
        IT_MemInt_header = self.td_single(config_file, 'IT_MemInt_Header')
        IT_MemInt_data = self.td_multiple(config_file, 'IT_MemInt_Tag', sheet, start_index=s.MOTOR_START_INDEX)

        if IT_IOInt_data != '' and IT_IOInt_header != '' and IT_MemInt_header != '' and IT_MemInt_data != '':
            filename = 'IT_' + sheet + '.csv'
            file_and_path = os.path.join(file_path, filename)
            with open(file_and_path, 'w') as itFile:
                data = IT_IOInt_header
                data += IT_IOInt_data
                data += IT_MemInt_header
                data += IT_MemInt_data

                itFile.write(data)
                print(filename, 'created')
                logging.info(filename + ' created')
        print('Generated files put in...', file_path)
        logging.info('Generated Motor files put in ' + file_path)

    def td_gen_AI(self):
        """Create and concetenate all text lines to different files"""
        # setup variables
        config_file = os.path.join(s.CONFIG_PATH, 'Config_AI.txt')
        sheet = 'AI'

        # Check what output path to use, if 'None' create in current directory, otherwise as specified
        if self.output_path is None:
            file_path = 'Generated AI'
        elif self.output_path == OUTPUT_PATH_START_VALUE:
            file_path = 'Generated AI'
        else:
            file_path = os.path.join(self.output_path, 'Generated AI')
        # Create sub-directory if it doesn't exist
        if not os.path.exists(file_path):
            os.makedirs(file_path)

        """Intouch IO:Int"""
        IT_IOInt_header = self.td_single(config_file, 'IT_IOInt_Header')
        IT_IOInt_data = self.td_multiple(config_file, 'IT_IOInt_Tag', sheet, udt_size=20, udt_offset=0,
                                         start_index=s.AI_START_INDEX)
        """Intouch Memory:Int"""
        IT_MemInt_header = self.td_single(config_file, 'IT_MemInt_Header')
        IT_MemInt_data = self.td_multiple(config_file, 'IT_MemInt_Tag', sheet, start_index=s.AI_START_INDEX)

        """Intouch IO:Real"""
        IT_IOReal_header = self.td_single(config_file, 'IT_IOReal_Header')
        IT_IOReal_data = self.td_multiple(config_file, 'IT_IOReal_Tag', sheet, udt_size=20, udt_offset=16,
                                          start_index=s.AI_START_INDEX)

        if IT_IOInt_data != '' and IT_IOInt_header != '' and IT_MemInt_header != '' and IT_MemInt_data != '':
            filename = 'IT_' + sheet + '.csv'
            file_and_path = os.path.join(file_path, filename)
            with open(file_and_path, 'w') as itFile:
                data = IT_IOInt_header
                data += IT_IOInt_data
                data += IT_MemInt_header
                data += IT_MemInt_data
                data += IT_IOReal_header
                data += IT_IOReal_data
                itFile.write(data)
                print(filename, 'created')
                logging.info(filename + ' created')
        print('Generated files put in...', file_path)
        logging.info('Generated AI files put in ' + file_path)


# Call UI
root = tk.Tk()
app = TdUI(master=root)
app.mainloop()

# Dump user data
with open('user_data.pickle', 'wb') as f:
    pickle.dump(user_data, f)
