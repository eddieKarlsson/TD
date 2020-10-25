import os
import sys
import os.path
import openpyxl as xl
from settings import Settings


class GenEngine:
    """Multiple functions to interact with worbook and sub-classes in obj_lib,
    and ui.
    """
    def __init__(self, excel_path, output_path):
        self.excel_path = excel_path
        self.output_path = output_path
        self.s = Settings()
        self.all_it_files = []  # Create an empty list
        self.dict_list = []

        self.generate()

    def _open_gen_excel(self):
        try:
            wb = xl.load_workbook(self.excel_path, data_only=True)
        except FileNotFoundError as e:
            print(e)
            print('Error! Excel file not found, program will exit')
            sys.exit()
        else:
            self.wb = wb

    def copy_excel_data_to_dictionaries(self):
        """Open excel and read all relevant object-data to dict"""
        self._open_gen_excel()

        # Create all dictionaries, if enabled in settings
        if not self.s.DI_DISABLE:
            self.di_dict = self._obj_data_to_dict(
                        self.s.DI_SHEETNAME, self.s.DI_START_INDEX, 'di')
            self.dict_list.append(self.di_dict)
        if not self.s.DO_DISABLE:
            self.do_dict = self._obj_data_to_dict(
                        self.s.DO_SHEETNAME, self.s.DO_START_INDEX, 'do')
            self.dict_list.append(self.do_dict)
        if not self.s.VALVE_DISABLE:
            self.valve_dict = self._obj_data_to_dict(
                self.s.VALVE_SHEETNAME, self.s.VALVE_START_INDEX, 'valve',
                config=True)
            self.dict_list.append(self.valve_dict)
        if not self.s.MOTOR_DISABLE:
            self.motor_dict = self._obj_data_to_dict(
                            self.s.MOTOR_SHEETNAME, self.s.MOTOR_START_INDEX,
                            'motor')
            self.dict_list.append(self.motor_dict)
        if not self.s.AI_DISABLE:
            self.ai_dict = self._obj_data_to_dict(
                self.s.AI_SHEETNAME, self.s.AI_START_INDEX, 'ai',
                eng_var=True)
            self.dict_list.append(self.ai_dict)
        if not self.s.AO_DISABLE:
            self.ao_dict = self._obj_data_to_dict(
                    self.s.AO_SHEETNAME, self.s.AO_START_INDEX, 'ao',
                    eng_var=True)
            self.dict_list.append(self.ao_dict)

    def _obj_data_to_dict(self, sheet, start_index, type,
                          config=False, eng_var=False):
        """Read all object data to dict"""

        # Open excel sheet
        try:
            ws = self.wb[sheet]
        except KeyError:
            msg = f'Error! {sheet} sheet does not exist, prog will exit'
            print(msg)
            sys.exit()

        # Set column indexes, automatically or hard-coded
        if self.s.FIND_HEADERS_AUTOMATICALLY:
            # Loop header and set the corresponding variables to
            # the integer number
            for i in range(1, 20):
                cell = ws.cell(row=self.s.HEADER_ROW, column=i)
                cellval = str(cell.value)

                # If cell is empty (NoneType) skip it
                if cellval is None:
                    continue

                if self.s.COL_ID_NAME in cellval:
                    column_id = i
                if self.s.COL_COMMENT_NAME in cellval:
                    column_comment = i

                if config:
                    if self.s.COL_CONFIG_NAME in cellval:
                        column_config = i

                if eng_var:
                    if self.s.COL_ENG_UNIT_NAME in cellval:
                        column_eng_unit = i
                    if self.s.COL_ENG_MIN_NAME in cellval:
                        column_eng_min = i
                    if self.s.COL_ENG_MAX_NAME in cellval:
                        column_eng_max = i
        else:
            column_id = self.s.COL_ID
            column_comment = self.s.COL_COMMENT
            if config:
                column_config = self.s.COL_CONFIG
            if eng_var:
                column_eng_unit = self.s.COL_ENG_UNIT
                column_eng_min = self.s.COL_ENG_MIN
                column_eng_max = self.s.COL_ENG_MAX

        if self.s.debug_level > 0:
            print('SHEET:', sheet)
            print('\t', 'column_id:', column_id)
            print('\t', 'column_comment:', column_comment)
            if config:
                print('\t', 'column_config:', column_config)
            if eng_var:
                print('\t', 'column_eng_unit:', column_eng_unit)
                print('\t', 'column_eng_min:', column_eng_min)
                print('\t', 'column_eng_max:', column_eng_max)

        # Loop through object list and add key-value pairs to object dict
        # then append each object-dict to list
        obj_list = []
        idx = start_index
        for i in range(self.s.ROW, ws.max_row + 1):
            # Break if we get a blank ID cell
            cell_id = ws.cell(row=i, column=column_id)
            cell_comment = ws.cell(row=i, column=column_comment)
            if cell_id.value is None:
                break

            # Always insert these key-value pairs
            obj = {
                'type': type,
                'id': cell_id.value,
                'comment': cell_comment.value,
                'index': idx,
            }

            # Add conditional key-value pairs
            if config:
                cell_config = ws.cell(row=i, column=column_config)
                obj['config'] = cell_config.value

            if eng_var:
                cell_eng_unit = ws.cell(row=i, column=column_eng_unit)
                obj['eng_unit'] = cell_eng_unit.value
                cell_eng_min = ws.cell(row=i, column=column_eng_min)
                obj['eng_min'] = cell_eng_min.value
                cell_eng_max = ws.cell(row=i, column=column_eng_max)
                obj['eng_max'] = cell_eng_max.value

            obj_list.append(obj)
            idx += 1

        return obj_list

    def _replace_keywords(self, line, obj, data_size, data_offset):
        """Take in a line and convert all the identifiers to obj data"""

        # Replace the keywords that always exists
        line = line.replace(self.s.ID_REPLACE, obj['id'])

        # check if comment exists, if not insert empty string
        if obj['comment'] is None:
            line = line.replace(self.s.COMMENT_REPLACE, '')
        else:
            line = line.replace(self.s.COMMENT_REPLACE,
                                obj['comment'])

        # Replace index
        line = line.replace(self.s.INDEX_REPLACE, obj['index'])

        # calculate address by offset & datatype data_size
        adress = (obj['index'] * data_size) + data_offset
        # Replace '@ADR'
        line = line.replace(self.s.ADR_REPLACE, str(adress))

        # Replace PLC
        line = line.replace(self.s.PLC_REPLACE,
                            self.s.PLC_NAME)

        # Replace the keywords that are optional (check if they exist)

        if obj.get('config') is not None:
            line = line.replace(self.s.CONFIG_REPLACE, obj['config'])

        if obj.get('eng_unit') is not None:
            line = line.replace(self.s.ENG_UNIT_REPLACE,
                                obj['eng_unit')
        else:
            line = line.replace(self.s.ENG_UNIT_REPLACE, '')

        if obj.get('eng_min') is not None:
            line = line.replace(self.s.ENG_MIN_REPLACE,
                                obj['eng_min')
        else:
            line = line.replace(self.s.ENG_MIN_REPLACE, '0')

        if obj.get('eng_max') is not None:
            line = line.replace(self.s.ENG_MAX_REPLACE,
                                obj['eng_max')
        else:
            line = line.replace(self.s.ENG_MAX_REPLACE, '100')

        return line

    @staticmethod
    def single(config_file, ref_txt):
        """Read a text file and copy the data inside notifiers to memory"""
        with open(config_file, 'r') as config:
            exists_in_config = False
            section_found = False
            inst_data = ''
            begin = '[gen.' + ref_txt + '_begin]'
            end = '[gen.' + ref_txt + '_end]'

            for line in config:
                if end in str(line):
                    section_found = False
                if section_found:
                    inst_data += line
                if begin in str(line):
                    exists_in_config = True
                    section_found = True
        if not exists_in_config:
            resultOk = False
            resultMsg = 'ref_txt not found in config file'
        else:
            resultOk = True
            resultMsg = None

        # Return a dictionary with the result
        result = {
            'ref_txt': ref_txt,
            'excelsheet': None,
            'from_row': None,
            'to_row': None,
            'resultOk': resultOk,
            'badResultMsg': resultMsg
        }

        return inst_data, result

    def multiple(self, obj_list, config_file, ref_txt, excelsheet,
                 data_size=30, data_offset=14):
        """Get text lines from config file and replace by data in excel for
            each item, then append the new lines to memory"""

        with open(config_file, 'r') as config:
            exists_in_config = False
            section_found = False
            inst_data = ''
            begin = '[gen.' + ref_txt + '_begin]'
            end = '[gen.' + ref_txt + '_end]'

            for obj in obj_list:
                config.seek(0, 0)  # Seek to beginning of file
                for lineIndex, line in enumerate(config, start=1):
                    if end in str(line):
                        section_found = False
                    if section_found:
                        line = self._replace_keywords(line, obj,
                                                      data_size, data_offset)
                        inst_data += line
                    if begin in str(line):
                        exists_in_config = True
                        section_found = True
            if not exists_in_config:
                resultOk = False
                resultMsg = 'ref_txt not found in config file'
            else:
                resultOk = True
                resultMsg = None

            # Return a dictionary with the result
            result = {
                'ref_txt': ref_txt,
                'excelsheet': active_sheet,
                'from_row': self.s.row,
                'to_row': i - 1,
                'resultOk': resultOk,
                'badResultMsg': resultMsg
            }


        return inst_data, result

    def multiple_config(self, sub_dir, ref_txt, excelsheet):
        """Same as td_multiple, but config stored in different files"""
        # Try to open excelsheet, otherwise prompt user
        try:
            active_sheet = self.wb[excelsheet]  # open sheet
        except KeyError:
            msg = f'Error! {excelsheet} sheet does not exist, prog will exit'
            print(msg)
            sys.exit()

        # Setup variables
        exists_in_config = False
        section_found = False
        inst_data = ''
        begin = '[gen.' + ref_txt + '_begin]'
        end = '[gen.' + ref_txt + '_end]'
        zeroIdx = 0

        # loop through excel rows, get value at corresponding cell
        for i in range(self.s.ROW, active_sheet.max_row + 1):
            cell_id = active_sheet.cell(row=i, column=self.s.COL_ID)
            cell_config = active_sheet.cell(row=i, column=self.s.COL_CONFIG)
            cell_comment = active_sheet.cell(row=i, column=self.s.COL_COMMENT)
            zeroIdx += 1

            if cell_id.value is None:
                break

            # combine file path and open corresponding file
            filename = cell_config.value + '.txt'
            file_and_path = os.path.join(sub_dir, filename)
            with open(file_and_path, 'r') as config:
                for line in config:
                    if end in str(line):
                        section_found = False
                    if section_found:
                        line = line.replace(self.s.ID_REPLACE, cell_id.value)

                        if cell_config.value is not None:
                            line = line.replace(self.s.CONFIG_REPLACE,
                                                cell_config.value)
                        line = line.replace(self.s.INDEX_REPLACE, str(zeroIdx))

                        # check if comment exists, if insert empty string
                        if cell_comment.value is None:
                            line = line.replace(self.s.COMMENT_REPLACE, '')
                        else:
                            line = line.replace(self.s.COMMENT_REPLACE,
                                                cell_comment.value)

                        inst_data += line  # Create instance data
                    if begin in str(line):
                        exists_in_config = True
                        section_found = True
        if not exists_in_config:
            print(ref_txt, 'in config file not found!')
        else:
            print('Generated from row:', self.s.ROW, 'to',
                  i - 1, 'of', ref_txt, 'in', active_sheet)
        return inst_data

    def generate(self):
        """Logging settings"""
        print('Version', self.s.version)

        self.copy_excel_data_to_dictionaries()

        if self.s.debug_level > 0:
            for dict in self.dict_list:
                for obj in dict:
                    print(obj)


"""
        # DI
        if self.s.DI_DISABLE:
            print('DI not generated, disabled in settings file')
        else:
            self.td_gen_di()

        # DO
        if self.s.DO_DISABLE:
            print('DO not generated, disabled in settings file')
        else:
            self.td_gen_do()

        # Valve
        if self.s.VALVE_DISABLE:
            print('Valve not generated, disabled in settings file')
        else:
            self.td_gen_valve()

        # Motor
        if self.s.MOTOR_DISABLE:
            print('Motor not generated, disabled in settings file')
        else:
            self.td_gen_motor()

        # AI
        if self.s.AI_DISABLE:
            print('AI not generated, disabled in settings file')
        else:
            self.td_gen_ai()

        # AO
        if self.s.AO_DISABLE:
            print('AO not generated, disabled in settings file')
        else:
            self.td_gen_ao()
"""
