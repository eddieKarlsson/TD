import sys
import os
import os.path
import openpyxl as xl
from settings import Settings


class GenTD:
    """Various functions to interact with excel workbook in TD format"""
    def __init__(self, excel_path, output_path):
        self.excel_path = excel_path
        self.output_path = output_path
        self.s = Settings()
        self.all_it_files = []  # Create an empty list
        self.dict_list = []

        self.generate()

    def _open_td_excel(self):
        try:
            wb = xl.load_workbook(self.excel_path, data_only=True)
        except FileNotFoundError as e:
            print(e)
            print('Error! Excel file not found, program will exit')
            sys.exit()
        else:
            self.wb = wb

    def td_copy_excel_data_to_dictionaries(self):
        """Open excel and read all relevant object-data to dict"""
        self._open_td_excel()

        # Create all dictionaries, if enabled in settings
        if not self.s.DI_DISABLE:
            self.di_dict = self._td_obj_data_to_dict(
                        self.s.DI_SHEETNAME, self.s.DI_START_INDEX, 'di')
            self.dict_list.append(self.di_dict)
        if not self.s.DO_DISABLE:
            self.do_dict = self._td_obj_data_to_dict(
                        self.s.DO_SHEETNAME, self.s.DO_START_INDEX, 'do')
            self.dict_list.append(self.do_dict)
        if not self.s.VALVE_DISABLE:
            self.valve_dict = self._td_obj_data_to_dict(
                self.s.VALVE_SHEETNAME, self.s.VALVE_START_INDEX, 'valve',
                config=True)
            self.dict_list.append(self.valve_dict)
        if not self.s.MOTOR_DISABLE:
            self.motor_dict = self._td_obj_data_to_dict(
                            self.s.MOTOR_SHEETNAME, self.s.MOTOR_START_INDEX,
                            'motor')
            self.dict_list.append(self.motor_dict)
        if not self.s.AI_DISABLE:
            self.ai_dict = self._td_obj_data_to_dict(
                self.s.AI_SHEETNAME, self.s.AI_START_INDEX, 'ai',
                eng_var=True)
            self.dict_list.append(self.ai_dict)
        if not self.s.AO_DISABLE:
            self.ao_dict = self._td_obj_data_to_dict(
                    self.s.AO_SHEETNAME, self.s.AO_START_INDEX, 'ao',
                    eng_var=True)
            self.dict_list.append(self.ao_dict)

    def _td_obj_data_to_dict(self, sheet, start_index, type,
                             config=False, eng_var=False):
        """Read all object data to dict"""

        # Open excel sheet
        try:
            ws = self.wb[sheet]
        except KeyError:
            msg = f'Error! {sheet} sheet does not exist, prog will exit'
            print(msg)
            sys.exit()

        # Loop through object list and add key-value pairs to object dict
        # then append each object-dict to list
        obj_list = []
        idx = start_index
        for i in range(self.s.ROW, ws.max_row + 1):
            # Break if we get a blank ID cell
            cell_id = ws.cell(row=i, column=self.s.COL_ID)
            cell_comment = ws.cell(row=i, column=self.s.COL_COMMENT)
            if cell_id.value is None:
                break

            # Always insert these key-value pairs
            obj = {
                'type': type,
                'id': cell_id.value,
                'comment': cell_comment.value,
                'index': idx,
            }

            # Add conditional key-value pairs
            if config:
                cell_config = ws.cell(row=i, column=self.s.COL_CONFIG)
                obj['config'] = cell_config.value

            if eng_var:
                cell_eng_unit = ws.cell(row=i, column=self.s.COL_ENG_UNIT)
                obj['eng_unit'] = cell_eng_unit.value
                cell_eng_min = ws.cell(row=i, column=self.s.COL_ENG_MIN)
                obj['eng_min'] = cell_eng_min.value
                cell_eng_max = ws.cell(row=i, column=self.s.COL_ENG_MAX)
                obj['eng_max'] = cell_eng_max.value

            obj_list.append(obj)
            idx += 1

        return obj_list

    def td_single(self, config_file, ref_txt):
        """Read a text file and copy the data inside notifiers to memory"""
        with open(config_file, 'r') as config:
            exists_in_config = False
            section_found = False
            inst_data = ''
            begin = '[gen.' + ref_txt + '_begin]'
            end = '[gen.' + ref_txt + '_end]'

            for line in config:
                if end in str(line):
                    section_found = False
                if section_found:
                    inst_data += line
                if begin in str(line):
                    exists_in_config = True
                    section_found = True
        if not exists_in_config:
            print(ref_txt, 'not found in config file!')

        return inst_data

    def td_multiple(self, config_file, ref_txt, excelsheet, udt_size=30,
                    udt_offset=14, start_index=0):
        """Get text lines from config file and replace by data in excel,
           then append the new lines to memory"""
        # Try to open excelsheet, otherwise prompt user
        try:
            active_sheet = self.wb[excelsheet]  # open sheet
        except KeyError:
            msg = f'Error! {excelsheet} sheet does not exist, prog will exit'
            print(msg)
            sys.exit()

        with open(config_file, 'r') as config:
            exists_in_config = False
            section_found = False
            inst_data = ''
            begin = '[gen.' + ref_txt + '_begin]'
            end = '[gen.' + ref_txt + '_end]'
            adrIndx = start_index - 1  # -1 to have first 0

            for i in range(self.s.ROW, active_sheet.max_row + 1):
                cell_id = active_sheet.cell(row=i, column=self.s.COL_ID)
                cell_config = active_sheet.cell(row=i,
                                                column=self.s.COL_CONFIG)
                cell_comment = active_sheet.cell(row=i,
                                                 column=self.s.COL_COMMENT)
                cell_engunit = active_sheet.cell(row=i,
                                                 column=self.s.COL_ENG_UNIT)
                cell_engmin = active_sheet.cell(row=i,
                                                column=self.s.COL_ENG_MIN)
                cell_engmax = active_sheet.cell(row=i,
                                                column=self.s.COL_ENG_MAX)

                adrIndx += 1

                if cell_id.value is None:
                    break

                config.seek(0, 0)  # Seek to beginning of file
                for lineIndex, line in enumerate(config, start=1):
                    if end in str(line):
                        section_found = False
                    if section_found:
                        line = line.replace(self.s.ID_REPLACE, cell_id.value)

                        if cell_config.value is not None:
                            line = line.replace(self.s.CONFIG_REPLACE,
                                                cell_config.value)

                        # Replace index
                        line = line.replace(self.s.INDEX_REPLACE, str(adrIndx))

                        # calculate address by offset & datatype udt_size
                        adress = (adrIndx * udt_size) + udt_offset
                        # Replace '@ADR'
                        line = line.replace(self.s.ADR_REPLACE, str(adress))

                        # Replace PLC
                        line = line.replace(self.s.PLC_REPLACE,
                                            self.s.PLC_NAME)

                        # check if eng unit exists, if not insert empty string
                        if cell_engunit.value is None:
                            line = line.replace(self.s.ENG_UNIT_REPLACE, '')
                        else:
                            line = line.replace(self.s.ENG_UNIT_REPLACE,
                                                cell_engunit.value)

                        # check if eng min exists, if not insert 0
                        if cell_engmin.value is None:
                            line = line.replace(self.s.ENG_MIN_REPLACE, '0')
                        else:
                            line = line.replace(self.s.ENG_MIN_REPLACE,
                                                str(cell_engmin.value))

                        # check if eng max exists, if not insert 100
                        if cell_engmax.value is None:
                            line = line.replace(self.s.ENG_MAX_REPLACE, '100')
                        else:
                            line = line.replace(self.s.ENG_MAX_REPLACE,
                                                str(cell_engmax.value))

                        # check if comment exists, if not insert empty string
                        if cell_comment.value is None:
                            line = line.replace(self.s.COMMENT_REPLACE, '')
                        else:
                            line = line.replace(self.s.COMMENT_REPLACE,
                                                cell_comment.value)

                        inst_data += line  # Create instance data
                    if begin in str(line):
                        exists_in_config = True
                        section_found = True
            if not exists_in_config:
                print(ref_txt, 'in config file not found!')
            else:
                print('Generated from row:',
                      self.s.ROW, 'to', i - 1, 'of', ref_txt, 'in',
                      active_sheet)

        return inst_data

    def td_multiple_config(self, sub_dir, ref_txt, excelsheet):
        """Same as td_multiple, but config stored in different files"""
        # Try to open excelsheet, otherwise prompt user
        try:
            active_sheet = self.wb[excelsheet]  # open sheet
        except KeyError:
            msg = f'Error! {excelsheet} sheet does not exist, prog will exit'
            print(msg)
            sys.exit()

        # Setup variables
        exists_in_config = False
        section_found = False
        inst_data = ''
        begin = '[gen.' + ref_txt + '_begin]'
        end = '[gen.' + ref_txt + '_end]'
        zeroIdx = 0

        # loop through excel rows, get value at corresponding cell
        for i in range(self.s.ROW, active_sheet.max_row + 1):
            cell_id = active_sheet.cell(row=i, column=self.s.COL_ID)
            cell_config = active_sheet.cell(row=i, column=self.s.COL_CONFIG)
            cell_comment = active_sheet.cell(row=i, column=self.s.COL_COMMENT)
            zeroIdx += 1

            if cell_id.value is None:
                break

            # combine file path and open corresponding file
            filename = cell_config.value + '.txt'
            file_and_path = os.path.join(sub_dir, filename)
            with open(file_and_path, 'r') as config:
                for line in config:
                    if end in str(line):
                        section_found = False
                    if section_found:
                        line = line.replace(self.s.ID_REPLACE, cell_id.value)

                        if cell_config.value is not None:
                            line = line.replace(self.s.CONFIG_REPLACE,
                                                cell_config.value)
                        line = line.replace(self.s.INDEX_REPLACE, str(zeroIdx))

                        # check if comment exists, if insert empty string
                        if cell_comment.value is None:
                            line = line.replace(self.s.COMMENT_REPLACE, '')
                        else:
                            line = line.replace(self.s.COMMENT_REPLACE,
                                                cell_comment.value)

                        inst_data += line  # Create instance data
                    if begin in str(line):
                        exists_in_config = True
                        section_found = True
        if not exists_in_config:
            print(ref_txt, 'in config file not found!')
        else:
            print('Generated from row:', self.s.ROW, 'to',
                  i - 1, 'of', ref_txt, 'in', active_sheet)
        return inst_data

    def generate(self):
        """Logging settings"""
        print('Version', self.s.version)

        self.td_copy_excel_data_to_dictionaries()

        if self.s.debug_level > 0:
            for dict in self.dict_list:
                for obj in dict:
                    print(obj)


"""
        # DI
        if self.s.DI_DISABLE:
            print('DI not generated, disabled in settings file')
        else:
            self.td_gen_di()

        # DO
        if self.s.DO_DISABLE:
            print('DO not generated, disabled in settings file')
        else:
            self.td_gen_do()

        # Valve
        if self.s.VALVE_DISABLE:
            print('Valve not generated, disabled in settings file')
        else:
            self.td_gen_valve()

        # Motor
        if self.s.MOTOR_DISABLE:
            print('Motor not generated, disabled in settings file')
        else:
            self.td_gen_motor()

        # AI
        if self.s.AI_DISABLE:
            print('AI not generated, disabled in settings file')
        else:
            self.td_gen_ai()

        # AO
        if self.s.AO_DISABLE:
            print('AO not generated, disabled in settings file')
        else:
            self.td_gen_ao()
"""
